# coding: utf-8
import os
import random
import shutil
from datetime import date, datetime, timedelta
import calendar
import pdfplumber
from openpyxl import load_workbook, Workbook
from openpyxl.styles import NamedStyle, Font, Border, Side, PatternFill
from xml.dom.minidom import Document
from pyWebFramework.core.task import TaskBase
from pyWebFramework import FailureException, SuccessException
from pyWebFramework.dll.IeWebFramework import TaxInitApi
from pyWebFramework.tax_init.PageTaxInitFormBase import PageTaxInitFormBase
from pyWebFramework.core.settings import settings


def pdf_to_xlsx(pdf_file, xlsx_file):
    workbook = Workbook()
    workbook.remove_sheet(workbook.active)

    has_sheet = False

    pdf = pdfplumber.open(pdf_file)
    sheet_index = 1
    for page in pdf.pages:
        for table in page.extract_tables({
        'vertical_strategy': 'lines',
        'horizontal_strategy': 'lines',
        'explicit_vertical_lines': page.curves + page.edges,
        'explicit_horizontal_lines': page.curves + page.edges,
    }):
            has_sheet = True
            sheet = workbook.create_sheet('Sheet%d' % sheet_index)
            line = 0
            for row in table:
                for j in range(len(row)):
                    sheet.cell(row=line + 1, column=j + 1).value = row[j]
                line += 1
            sheet_index += 1

    pdf.close()

    if has_sheet:
        workbook.save(xlsx_file)

    return has_sheet


def combine_xlsxs(xlsx_files, target_xlsx):
    target_wb = Workbook()
    target_wb.remove_sheet(target_wb.active)

    sheet_index = 1
    for xlsx_file in xlsx_files:
        wb = load_workbook(xlsx_file)
        for sheet_name in wb.get_sheet_names():
            sheet = wb.get_sheet_by_name(sheet_name)
            target_sheet = target_wb.create_sheet('Sheet%d' % sheet_index)
            sheet_index += 1
            for row in sheet.iter_rows():
                for cell in row:
                    target_sheet[cell.coordinate] = cell.value

    target_wb.save(target_xlsx)


class TaskTaxInitBase(TaskBase):
    def __init__(self):
        super(TaskTaxInitBase, self).__init__()

        self.tables = []
        self.__current_table_index = 0

        self.samples = []

        self.collected_data = []

        self.clean_template_files()

    def InitTask(self):
        if not super(TaskTaxInitBase, self).InitTask():
            return False

        if not settings.TAX_INIT_SAMPLE_ROOT:
            self.Failure('没有配置 TAX_INIT_SAMPLE_ROOT')
            return False

        if not self.LoadSamples(settings.TAX_INIT_SAMPLE_ROOT):
            return False

        for table_id in self.GetTables():
            table = {
                'id': table_id,
                'name': self.GetTableAttrById(table_id, 'name'),
                'sbqs': self.GetTableAttrById(table_id, 'sbqs'),
                'sbqz': self.GetTableAttrById(table_id, 'sbqz'),
                'ssqs': self.GetTableAttrById(table_id, 'ssqs'),
                'ssqz': self.GetTableAttrById(table_id, 'ssqz'),
                'type': self.GetTableAttrById(table_id, 'type'),
                'ssqType': self.GetTableAttrById(table_id, 'ssqType'),
            }
            table['sample'] = self.SelectTable(table)
            self.tables.append(table)

        if not self.tables:
            raise FailureException('没有表')

        return True

    def BeginTask(self):
        if not super(TaskTaxInitBase, self).BeginTask():
            return False

        return True

    def Success(self, desc=''):
        xml = self.GenerateXmlString()
        open(os.path.join(settings.TAX_INIT_SAMPLE_ROOT, 'result.xml'), 'w', encoding='utf-8').write(xml)

        last_success_msg = ''
        for tableset_data in self.collected_data:
            tableset_xml = self.GenerateXmlString(tableset_data)

            api = TaxInitApi(False)
            result = api.request(tableset_xml)

            if not result.success:
                self.Failure('调用接口失败: code=%s msg=%s' % (result.code, result.msg))
                return

            last_success_msg = result.msg

        super(TaskTaxInitBase, self).Success('调用接口成功，' + last_success_msg)

    def LoadSamples(self, sample_root):
        for file in os.listdir(sample_root):
            file_no_ext, file_ext = os.path.splitext(file)
            if file.lower().endswith('.cache.xlsx'):
                continue
            if file.lower().endswith('.changes.xlsx'):
                continue
            if file_ext.lower() != '.xlsx':
                continue
            if file_no_ext.startswith('~$'):
                continue
            file_sep = file_no_ext.split('_')
            if len(file_sep) < 3:
                continue

            table_id = file_sep[-3]
            sbqq = file_sep[-2]
            sbqz = file_sep[-1]

            if len(sbqq) is not 6:
                continue

            if sbqz != 'now' and len(sbqz) is not 6:
                continue

            sbqq = date(year=int(sbqq[:4]), month=int(sbqq[4:6]), day=1)

            def last_day_of_month(any_day):
                next_month = any_day.replace(day=28) + timedelta(days=4)  # this will never fail
                last_day = next_month - timedelta(days=next_month.day)

                return last_day.day

            if sbqz == 'now':
                now = datetime.now()
                last_day = last_day_of_month(now)
                sbqz = date(year=now.year, month=now.month, day=last_day)
            else:
                year = int(sbqz[:4])
                month = int(sbqz[4:6])
                day = calendar.monthrange(year, month)[1]
                sbqz = date(year=year, month=month, day=day)

            self.samples.append({
                'table_id': table_id,
                'name': file_sep[0],
                'sbqq': sbqq,
                'sbqz': sbqz,
                'sample_file': os.path.join(sample_root, file),
                'cache_file': os.path.join(sample_root, file_no_ext + '.cache.xlsx'),
                'changes_file': os.path.join(sample_root, file_no_ext + '.changes.xlsx'),
            })
        return True

    def GetNextTable(self):
        if self.__current_table_index >= len(self.tables):
            return None

        table = self.tables[self.__current_table_index]

        self.__current_table_index += 1

        return table

    def SelectTable(self, table):
        table_id = table['id']
        sbqq = table['sbqs']
        sbqz = table['sbqz']

        sbqq = sbqq .replace('-', '')
        sbqz = sbqz .replace('-', '')
        if len(sbqz) is not 8 or len(sbqz) is not 8:
            return None

        sbqq = date(year=int(sbqq[:4]), month=int(sbqq[4:6]), day=int(sbqq[6:8]))
        sbqz = date(year=int(sbqz[:4]), month=int(sbqz[4:6]), day=int(sbqz[6:8]))

        for sample in self.samples:
            if sample['table_id'] != table_id:
                continue

            sample_sbqq = sample['sbqq']
            sample_sbqz = sample['sbqz']

            if sbqq >= sample_sbqq and sbqz <= sample_sbqz:
                return sample

        return None

    def SelectSampleFile(self, table):
        obj = self.SelectTable(table)
        if not obj:
            return None
        return obj['sample_file']

    def SelectCacheFile(self, table):
        obj = self.SelectTable(table)
        if not obj:
            return None
        return obj['cache_file']

    def SelectChangesFile(self, table):
        obj = self.SelectTable(table)
        if not obj:
            return None
        return obj['changes_file']

    def LoadSampleCells(self, sample_file):
        keywords = ('cell', 'label', 'verify', 'float', 'floatconfig', 'sheet', 'combine_sheet')

        datas = {}
        workbook = load_workbook(sample_file)
        for sheet_name in workbook.get_sheet_names():
            sheet = workbook.get_sheet_by_name(sheet_name)
            sheet_data = {
                'cells': [],
                'labels': [],
                'floats': {},
                'verifies': [],
                'sheet': [],
                'combine_sheet': [],
            }
            for row in sheet.iter_rows():
                for cell in row:
                    if not cell.value:
                        continue

                    cell_text = str(cell.value)

                    sep = cell_text.split('&')

                    column_letter = cell.column
                    if 'column_letter' in cell.__dir__():
                        column_letter = cell.column_letter

                    cell_data = {
                        'coordinate': cell.coordinate,
                        'row': cell.row,
                        'column': cell.col_idx,
                        'column_letter': column_letter,
                    }

                    if not sep:
                        continue

                    if len(sep) < 2:
                        continue

                    if sep[0].lower() not in keywords and sep[1].lower() in keywords:
                        comment = sep[0]
                        sep = sep[1:]
                        cell_data['comment'] = comment

                    cell_type = sep[0].lower()

                    if cell_type not in keywords:
                        continue

                    if cell_type == 'verify':
                        cell_data['expr'] = sep[1]
                        sheet_data['verifies'].append(cell_data)
                        continue

                    for attr in sep[1:]:
                        attr_sep = attr.split('=')
                        if len(attr_sep) < 2:
                            continue
                        key = attr_sep[0].strip()
                        value = attr_sep[1].strip()
                        if cell_type in ['sheet', 'combine_sheet']:
                            if 'cells' not in cell_data:
                                cell_data['cells'] = {}
                            cell_data['cells'][key] = value
                        else:
                            cell_data[key] = value

                    if cell_type == 'cell':
                        sheet_data['cells'].append(cell_data)
                    elif cell_type == 'label':
                        sheet_data['labels'].append(cell_data)
                    elif cell_type == 'float':
                        if 'name' not in cell_data:
                            continue
                        cell_name = cell_data['name']
                        if cell_name in sheet_data['floats']:
                            sheet_data['floats'][cell_name].append(cell_data)
                        else:
                            sheet_data['floats'][cell_name] = [cell_data]
                    elif cell_type == 'floatconfig':
                        if 'name' not in cell_data:
                            continue
                        cell_name = cell_data['name']
                        cell_name += '_config'
                        sheet_data['floats'][cell_name] = cell_data
                    elif cell_type == 'sheet':
                        sheet_data['sheet'] = cell_data
                    elif cell_type == 'combine_sheet':
                        sheet_data['combine_sheet'] = cell_data

            datas[sheet_name] = sheet_data

        return datas

    def CreateCacheFile(self, sample_file, cache_file):
        shutil.copy2(sample_file, cache_file)

        sample_cells = self.LoadSampleCells(cache_file)

        wb = load_workbook(cache_file)
        for sheet_name in sample_cells:
            sheet = wb.get_sheet_by_name(sheet_name)
            cells = sample_cells[sheet_name]
            if cells['sheet']:
                sheet[cells['sheet']['coordinate']].value = ''
            if cells['combine_sheet']:
                sheet[cells['combine_sheet']['coordinate']].value = ''
            for verify in cells['verifies']:
                sheet[verify['coordinate']].value = ''
            for label in cells['labels']:
                sheet[label['coordinate']].value = ''
            for cell in cells['cells']:
                sheet[cell['coordinate']].value = ''
            for float_name in cells['floats']:
                if float_name.endswith('_config'):
                    float_datas = cells['floats'][float_name]
                    sheet[float_datas['coordinate']].value = ''
                else:
                    float_datas = cells['floats'][float_name]
                    for float_data in float_datas:
                        sheet[float_data['coordinate']].value = ''

        wb.save(cache_file)

    def Collect(self, target, table):
        if isinstance(target, PageTaxInitFormBase):
            handler = self.CollectPage
        elif type(target) == str and target.lower().endswith('.pdf'):
            pdf_file = target
            dir_path = os.path.dirname(pdf_file)
            basename = os.path.basename(pdf_file)
            xlsx_file = os.path.join(dir_path, os.path.splitext(basename)[0] + '.xlsx')

            if not pdf_to_xlsx(pdf_file, xlsx_file):
                raise FailureException('pdf转excel失败')

            target = load_workbook(xlsx_file)

            handler = self.CollectExcel
        elif type(target) == list and target and target[0].lower().endswith('.pdf'):
            xlsx_files = []
            target_xlsx = ''
            for pdf_file in target:
                dir_path = os.path.dirname(pdf_file)
                basename = os.path.basename(pdf_file)
                xlsx_file = os.path.join(dir_path, os.path.splitext(basename)[0] + '.xlsx')

                if not target_xlsx:
                    target_xlsx = os.path.join(dir_path, 'forms.xlsx')

                if not pdf_to_xlsx(pdf_file, xlsx_file):
                    continue

                xlsx_files.append(xlsx_file)

            if not xlsx_files:
                raise FailureException('pdf转excel失败')

            combine_xlsxs(xlsx_files, target_xlsx)

            target = load_workbook(target_xlsx)
            handler = self.CollectExcel
        elif type(target) == str and target.lower().endswith('.xlsx'):
            target = load_workbook(target)
            handler = self.CollectExcel
        elif type(target) == list and target and target[0].lower().endswith('.xlsx'):
            dir_path = os.path.dirname(target[0])
            target_xlsx = os.path.join(dir_path, 'forms.xlsx')

            combine_xlsxs(target, target_xlsx)

            target = load_workbook(target_xlsx)
            handler = self.CollectExcel
        else:
            raise TypeError('Unknown target: ' + target)

        sample_file = self.SelectSampleFile(table)
        cache_file = self.SelectCacheFile(table)
        changes_file = self.SelectChangesFile(table)

        if not sample_file:
            raise FailureException(
                '未找到样本文件 table_id=%s ssqq=%s ssqz=%s' % (table['id'], table['ssqs'], table['ssqz']))

        if not os.path.exists(cache_file):
            self.CreateCacheFile(sample_file, cache_file)

        sample_cells = self.LoadSampleCells(sample_file)

        cache_wb = load_workbook(cache_file)

        def is_sheet_empty(sheet):
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value:
                        return False
            return True

        cache_changed = False
        changed = False
        for table_name in sample_cells:
            table_cells = sample_cells[table_name]
            cache_sheet = cache_wb.get_sheet_by_name(table_name)
            new_cache = is_sheet_empty(cache_sheet)
            if new_cache:
                cache_changed = True
            if not handler(target, cache_sheet, new_cache, table, table_name, table_cells):
                changed = True

        if cache_changed:
            cache_wb.save(cache_file)
        if changed:
            cache_wb.save(changes_file)
            raise FailureException('表格发生变化，table_id=%s ssqq=%s ssqz=%s，查看excel了解详细变化: %s' % (table['id'], table['ssqs'], table['ssqz'], changes_file))

    def CollectExcel(self, workbook, cache_sheet, new_cache, table, table_name, cells):
        if 'sheet' not in cells or 'cells' not in cells['sheet']:
            raise FailureException('没有配置sheet')

        def find_sheet(cells):
            for sheet_name in workbook.get_sheet_names():
                _sheet = workbook.get_sheet_by_name(sheet_name)

                found = True
                for key in cells:
                    value = cells[key]
                    _sheet_value = _sheet[key].value
                    if _sheet_value and type(_sheet_value) == str:
                        _sheet_value = _sheet_value.replace('\r\n', '')
                        _sheet_value = _sheet_value.replace('\n', '')
                    if _sheet_value != value:
                        found = False
                        break

                if found:
                    return _sheet
            return None

        sheet = find_sheet(cells['sheet']['cells'])

        if not sheet:
            # raise FailureException('没有找到sheet')
            return True

        if 'combine_sheet' in cells and 'cells' in cells['combine_sheet'] and cells['combine_sheet']['cells']:
            target_sheet = find_sheet(cells['combine_sheet']['cells'])
            if target_sheet:
                max_row = 0
                for row in sheet.iter_rows():
                    for cell in row:
                        if not cell.value:
                            continue
                        if cell.row > max_row:
                            max_row = cell.row

                for row in target_sheet.iter_rows():
                    for cell in row:
                        sheet.cell(max_row + cell.row, cell.column).value = cell.value

                workbook.remove(target_sheet)

        def reset_below_cells_row(base_row, count):
            def reset_cell_row(cell, offset):
                if 'cache_row' not in cell:
                    cell['cache_row'] = cell['row']
                cell['row'] += offset
                digit_idx = -1
                coordinate = cell['coordinate']
                for index, l in enumerate(coordinate):
                    if l.isdigit():
                        digit_idx = index
                        break
                col_letter = coordinate[:digit_idx]
                row = int(coordinate[digit_idx:]) + offset
                if 'cache_coordinate' not in cell:
                    cell['cache_coordinate'] = cell['coordinate']
                cell['coordinate'] = '%s%d' % (col_letter, row)
            row_offset = count - 1
            for float_name in cells['floats']:
                if float_name.endswith('_config'):
                    continue
                if float_name + '_config' not in cells['floats']:
                    continue
                float_datas = cells['floats'][float_name]
                for float_data in float_datas:
                    if float_data['row'] > base_row:
                        reset_cell_row(float_data, row_offset)

            for label in cells['labels']:
                if label['row'] > base_row:
                    reset_cell_row(label, row_offset)

            for cell in cells['cells']:
                if cell['row'] > base_row:
                    reset_cell_row(cell, row_offset)

        params = []

        float_row_num = 1
        for float_name in cells['floats']:
            if float_name.endswith('_config'):
                continue
            if float_name + '_config' not in cells['floats']:
                continue
            float_datas = cells['floats'][float_name]
            float_config = cells['floats'][float_name + '_config']

            id_base = int(float_config['id_base'])
            end_col = float_config['end_col']
            end_text = float_config['end_text']
            float_row_index = 0
            float_row_count = 0
            base_row = 0

            float_items = []
            params.append({
                'type': 'float',
                'name': float_name,
                'row_num': float_row_num,
                'data': float_items,
            })

            float_end = False
            while not float_end:
                for float_data in float_datas:
                    base_row = float_data['row']

                    _id = float_data['id'] % id_base
                    float_item_data = {
                        'id': _id,
                    }

                    row = float_data['row'] + float_row_index

                    float_item_data['value'] = sheet.cell(row=row, column=float_data['column']).value

                    if float_data['column_letter'] == end_col:
                        if (end_text and float_item_data['value'] == end_text) or (not end_text and not float_item_data['value']):
                            float_end = True
                            break

                    float_items.append(float_item_data)

                if not float_end:
                    id_base += 1
                    float_row_index += 1
                    float_row_count += 1

            reset_below_cells_row(base_row, float_row_count)
            float_row_num += 1

        verify_pass = True
        changed_fill = PatternFill("solid", fgColor="FF1111")
        for label in cells['labels']:
            comment = label['comment']
            if not comment:
                comment = '文字'

            cache_coordinate = label['coordinate']
            if 'cache_coordinate' in label:
                cache_coordinate = label['cache_coordinate']
            value = sheet[label['coordinate']].value

            cache_cell = cache_sheet[cache_coordinate]
            cache_cell.fill = PatternFill()

            if value:
                value = value.replace('\r\n', '\n')
            else:
                print('none')

            if new_cache:
                cache_cell.value = value
            elif not cache_cell.value:
                cache_cell.fill = changed_fill
                verify_pass = False
            elif not value:
                cache_cell.fill = changed_fill
                verify_pass = False
            else:
                if cache_cell.value.strip() != value.strip():
                    cache_cell.value = value
                    cache_cell.fill = changed_fill
                    verify_pass = False

        if not verify_pass:
            return False

        for cell in cells['cells']:
            coordinate = cell['coordinate']
            if 'orig_coordinate' in cell:
                coordinate = cell['orig_coordinate']

            value = sheet[coordinate].value
            if not value:
                value = ''

            if value == '--':
                value = ''

            value = value.replace('\r\n', '')
            value = value.replace('\n', '')

            # 数据中去除,
            if value and value[0].isdigit() and value[-1].isdigit():
                value = value.replace(',', '')
            elif len(value) >= 2 and value[0] == '-' and value[1].isdigit() and value[-1].isdigit():
                value = value.replace(',', '')

            # 百分数转小数
            if value and value[-1] == '%' and value[:-1].replace('.', '').isdigit():
                val = value[:-1]
                if val.count('.') == 1:
                    val_f = val.split('.')[1]
                    float_len = len(val_f) + 2
                else:
                    float_len = len(val)
                float_val = float(val) / 100.0
                fmt = '%%.%df' % float_len
                value = fmt % float_val

            item_data = {
                'id': cell['id'],
                'value': value,
            }

            self.OnCollectCell(item_data)

            params.append({
                'type': 'item',
                'data': item_data,
            })

        self.SaveTableData(table['id'], 'SB', '', table['ssqs'], table['ssqz'], table['ssqType'], table_name, params)

        return True

    def CollectPage(self, page: PageTaxInitFormBase, cache_sheet, new_cache, table, table_name, cells):
        for verify in cells['verifies']:
            comment = verify['comment']
            expr = verify['expr']
            if not comment:
                comment = expr
            if not page.EvalScriptBool(expr):
                raise FailureException('执行校验 %s 不通过 table_id=%s ssqq=%s ssqz=%s' % (comment, table['id'], table['ssqs'], table['ssqz']))

        verify_pass = True
        changed_fill = PatternFill("solid", fgColor="FF1111")
        for label in cells['labels']:
            comment = label['comment']
            if not comment:
                comment = '文字'
            if 'xpath' in label:
                if not page.IsCellExistByXPath(label['xpath']):
                    raise FailureException('检测 %s 不存在 name=%s xpath=%s' % (comment, table['name'], label['xpath']))
                value = page.GetCellTextByXPath(label['xpath'])
            elif 'selector' in label:
                if not page.IsCellExistBySelector(label['selector']):
                    raise FailureException('检测 %s 不存在 name=%s selector=%s' % (comment, table['name'], label['selector']))
                value = page.GetCellTextBySelector(label['selector'])
            else:
                continue
            cache_cell = cache_sheet[label['coordinate']]
            cache_cell.fill = PatternFill()

            if value:
                value = value.replace('\r\n', '\n')

            if new_cache:
                cache_cell.value = value
            else:
                if cache_cell.value.strip() != value.strip():
                    cache_cell.value = value
                    cache_cell.fill = changed_fill
                    verify_pass = False

        if not verify_pass:
            return False

        params = []
        for cell in cells['cells']:
            item_data = {
                'id': cell['id'],
            }

            value = ''
            if 'xpath' in cell:
                if not page.IsCellExistByXPath(cell['xpath']):
                    raise FailureException('税表检测到变化，单元格不存在 name=%s xpath=%s' % (table['name'], cell['xpath']))
                value = page.GetCellTextByXPath(cell['xpath'])
            elif 'selector' in cell:
                if not page.IsCellExistBySelector(cell['selector']):
                    raise FailureException('税表检测到变化，单元格不存在 name=%s selector=%s' % (table['name'], cell['selector']))
                value = page.GetCellTextBySelector(cell['selector'])

            if value == '--':
                value = ''

            value = value.replace('\r\n', '')
            value = value.replace('\n', '')

            item_data['value'] = value

            params.append({
                'type': 'item',
                'data': item_data,
            })

        float_row_num = 1
        for float_name in cells['floats']:
            if float_name.endswith('_config'):
                continue
            if float_name + '_config' not in cells['floats']:
                continue
            float_datas = cells['floats'][float_name]
            float_config = cells['floats'][float_name + '_config']

            xpath_base = 0
            selector_base = 0
            id_base = int(float_config['id_base'])
            count_expr = float_config['count_expr']
            if 'selector_base' in float_config:
                selector_base = int(float_config['selector_base'])
            if 'xpath_base' in float_config:
                xpath_base = int(float_config['xpath_base'])

            count = page.EvalScriptInt(count_expr)

            float_items = []
            params.append({
                'type': 'float',
                'name': float_name,
                'row_num': float_row_num,
                'data': float_items,
            })
            for i in range(count):
                for float_data in float_datas:
                    _id = float_data['id'] % id_base
                    float_item_data = {
                        'id': _id,
                    }

                    value = ''
                    if 'xpath' in float_data:
                        _xpath = float_data['xpath'] % xpath_base
                        if not page.IsCellExistByXPath(_xpath):
                            raise FailureException('税表检测到变化，浮动行单元格不存在 行=%d name=%s xpath=%s' % (i, table['name'], _xpath))
                        value = page.GetCellTextByXPath(_xpath)
                    elif 'selector' in float_data:
                        _selector = float_data['selector'] % selector_base
                        if not page.IsCellExistBySelector(_selector):
                            raise FailureException('税表检测到变化，浮动行单元格不存在 行=%d name=%s selector=%s' % (i, table['name'], _selector))
                        value = page.GetCellTextBySelector(_selector)

                    if value == '--':
                        value = ''

                    float_item_data['value'] = value

                    float_items.append(float_item_data)

                id_base += 1
                xpath_base += 1

            float_row_num += 1

        self.SaveTableData(table['id'], 'SB', '', table['ssqs'], table['ssqz'], table['ssqType'], table_name, params)

        return True

    def OnCollectCell(self, data):
        pass

    def SaveTableData(self, tableset_id, tableset_type, tax_type, ssqs, ssqz, ssq_type, table_name, params):
        for exist_data in self.collected_data:
            if exist_data['tableset_id'] == tableset_id \
                    and exist_data['tableset_type'] == tableset_type \
                    and exist_data['tax_type'] == tax_type \
                    and exist_data['ssqs'] == ssqs \
                    and exist_data['ssqz'] == ssqz \
                    and exist_data['ssq_type'] == ssq_type:
                tableset_data = exist_data
                break
        else:
            tableset_data = {
                'tableset_id': tableset_id,
                'tableset_type': tableset_type,
                'tax_type': tax_type,
                'ssqs': ssqs,
                'ssqz': ssqz,
                'ssq_type': ssq_type,
                'tables': [],
            }
            self.collected_data.append(tableset_data)

        for table in tableset_data['tables']:
            if table['name'] == table_name:
                table_data = table
                break
        else:
            table_data = {
                'name': table_name,
                'params': params,
            }
            tableset_data['tables'].append(table_data)

        table_data['params'] = params

    def GenerateXmlString(self, speicfic_tableset_data=None):
        doc = Document()

        Root = doc.createElement('Root')
        doc.appendChild(Root)

        TaskSet = doc.createElement('TaskSet')
        Root.appendChild(TaskSet)

        OrgInfo = doc.createElement('OrgInfo')
        TaskSet.appendChild(OrgInfo)

        OrgName = doc.createElement('OrgName')
        OrgInfo.appendChild(OrgName)
        OrgNameText = doc.createTextNode(self.GetLoginInfo('CompanyName'))
        OrgName.appendChild(OrgNameText)

        OrgTaxNum = doc.createElement('OrgTaxNum')
        OrgInfo.appendChild(OrgTaxNum)
        OrgTaxNumText = doc.createTextNode(self.GetLoginInfo('LoginNsrsbh'))
        OrgTaxNum.appendChild(OrgTaxNumText)

        LoginInfo = doc.createElement('LoginInfo')
        LoginInfo.setAttribute('LoginName', '')
        LoginInfo.setAttribute('LoginPassword', '')
        TaskSet.appendChild(LoginInfo)

        TaxInfo = doc.createElement('TaxInfo')
        TaskSet.appendChild(TaxInfo)

        collected_data = self.collected_data
        if speicfic_tableset_data:
            collected_data = [speicfic_tableset_data]

        for tableset_data in collected_data:
            TableSet = doc.createElement('TableSet')
            TaxInfo.appendChild(TableSet)

            tableset_id = tableset_data['tableset_id']
            if tableset_id.startswith('700'):
                tableset_id = '700000'
            TableSet.setAttribute('id', tableset_id)
            TableSet.setAttribute('type', tableset_data['tableset_type'])
            TableSet.setAttribute('taxType', tableset_data['tax_type'])
            TableSet.setAttribute('ssqs', tableset_data['ssqs'])
            TableSet.setAttribute('ssqz', tableset_data['ssqz'])
            TableSet.setAttribute('ssqType', tableset_data['ssq_type'])

            all_tables = []

            for table_data in tableset_data['tables']:
                table_name = table_data['name']
                if table_name.endswith('}') and table_name.rfind('{') != -1:
                    table_name = table_name[:table_name.rfind('{')]

                all_tables.append(table_name)

                Table = doc.createElement('Table')
                TableSet.appendChild(Table)

                Table.setAttribute('name', table_name)

                Param = doc.createElement('Param')
                Table.appendChild(Param)

                for param in table_data['params']:
                    def handle_items(ParentNode, item_data):
                        Item = doc.createElement('Item')
                        ParentNode.appendChild(Item)

                        if 'id' in item_data:
                            item_id = item_data['id']
                            Item.setAttribute('location', item_id)

                        text = doc.createTextNode(item_data['value'])

                        Item.appendChild(text)
                    if param['type'] == 'item':
                        item_data = param['data']
                        handle_items(Param, item_data)
                    elif param['type'] == 'float':
                        Float = doc.createElement('Float')
                        Param.appendChild(Float)

                        Float.setAttribute('name', param['name'])
                        Float.setAttribute('rowNum', str(param['row_num']))

                        for item_data in param['data']:
                            handle_items(Float, item_data)

            Table = doc.createElement('Table')
            TableSet.appendChild(Table)

            Table.setAttribute('name', '必填报税表列表')

            Param = doc.createElement('Param')
            Table.appendChild(Param)

            for table_name in all_tables:
                Item = doc.createElement('Item')
                Param.appendChild(Item)

                Item.setAttribute('col', '无列标题')
                Item.setAttribute('row', '无行标题')

                text = doc.createTextNode(table_name)

                Item.appendChild(text)

        return doc.toprettyxml(indent='\t', encoding='UTF-8').decode('utf-8', 'ignore')

    def make_template_file(self, name='', ext=''):
        dir = settings.TEMPLATE_FILE_ROOT
        if not dir:
            dir = os.path.join(settings.ROOT, 'temp')
        if not os.path.exists(dir):
            os.mkdir(dir)
        specific_name = False
        if name:
            specific_name = True
        else:
            name = ''.join(random.sample('abcdefghijklmnopqrstuvwxyz', 15))
        if ext and not ext.startswith('.'):
            ext = '.' + ext
        filepath = os.path.join(dir, name + ext)

        if not specific_name and os.path.exists(filepath):
            return self.make_template_file(name, ext)
        return filepath

    def clean_template_files(self):
        dir = settings.TEMPLATE_FILE_ROOT
        if not dir:
            dir = os.path.join(settings.ROOT, 'temp')
        if not os.path.exists(dir):
            return
        for name in os.listdir(dir):
            path = os.path.join(dir, name)
            if os.path.isfile(path):
                os.remove(path)


if __name__ == '__main__':
    pdf_to_xlsx(r'Y:\Downloads\YBNSRZZS.pdf', r'Y:\Downloads\YBNSRZZS.xlsx')

    task = TaskTaxInitBase()
    task.LoadSamples(r'Y:\Documents\Works\PyProjects\FuJian\任务\税务初始化\样本文件')

    currentTable = {
        'name': '增值税纳税申报表（一般纳税人适用）',
        'id': '010100',
        'ssqType': '',
        'sbqs': '2019-06-01',
        'sbqz': '2019-06-30',
        'ssqs': '2019-05-01',
        'ssqz': '2019-05-31',
        'type': '2',
    }

    task.Collect(r'Y:\Downloads\forms.xlsx', currentTable)

    task.Success()
